"""
capture_plan.py
Lee el plan de capturas desde un archivo JSON o Excel (.xlsx).
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

from view_controller import DisplayType, ViewConfig, ViewType

logger = logging.getLogger(__name__)


def load_plan(path: str | Path) -> list[ViewConfig]:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Plan no encontrado: {path}")

    suffix = path.suffix.lower()
    if suffix == ".json":
        return _load_json(path)
    if suffix in (".xlsx", ".xlsm"):
        return _load_excel(path)
    if suffix == ".xls":
        raise ValueError("Formato .xls no soportado; usa .xlsx, .xlsm o .json")
    raise ValueError(f"Formato no soportado: '{suffix}'. Usa .json, .xlsx o .xlsm")


def _load_json(path: Path) -> list[ViewConfig]:
    with path.open(encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, list):
        raise ValueError("El JSON debe ser una lista de objetos de captura")
    return [_dict_to_config(item, i) for i, item in enumerate(data, start=1)]


def _enum_from_data(enum_cls, raw_value, default, field_name: str, index: int):
    if raw_value is None:
        return default
    if isinstance(raw_value, enum_cls):
        return raw_value

    if isinstance(raw_value, str):
        candidate = raw_value.strip()
        if not candidate:
            return default
        try:
            return enum_cls[candidate.upper()]
        except KeyError:
            pass
        raw_value = candidate

    try:
        return enum_cls(int(raw_value))
    except (TypeError, ValueError) as exc:
        valid = [item.name for item in enum_cls]
        raise ValueError(
            f"Entrada {index}: {field_name}='{raw_value}' inválido. Opciones: {valid}"
        ) from exc


def _load_excel(path: Path) -> list[ViewConfig]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    if "filename" not in headers:
        raise ValueError(
            f"El Excel debe tener al menos la columna 'filename'. Encontradas: {headers}"
        )

    configs: list[ViewConfig] = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None for value in row):
            continue
        row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        try:
            configs.append(_dict_to_config(row_dict, row_num))
        except Exception as exc:
            logger.warning("Fila %s ignorada: %s", row_num, exc)
    return configs


def _dict_to_config(data: dict, index: int) -> ViewConfig:
    def _str(key: str, default: str = "") -> str:
        value = data.get(key)
        return str(value).strip() if value is not None else default

    def _int(key: str, default: int = 0) -> int:
        value = data.get(key)
        try:
            return int(value) if value is not None else default
        except (TypeError, ValueError):
            return default

    def _float(key: str, default: float = 0.0) -> float:
        value = data.get(key)
        try:
            return float(value) if value is not None else default
        except (TypeError, ValueError):
            return default

    def _bool(key: str, default: bool = False) -> bool:
        value = data.get(key)
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "si", "sí", "yes"}

    filename = _str("filename")
    if not filename:
        raise ValueError(f"Entrada {index}: 'filename' es obligatorio")

    view_type = _enum_from_data(
        ViewType,
        data.get("view_type", "ISO_3D"),
        ViewType.ISO_3D,
        "view_type",
        index,
    )
    display_type = _enum_from_data(
        DisplayType,
        data.get("display_type", "GEOMETRY_ONLY"),
        DisplayType.GEOMETRY_ONLY,
        "display_type",
        index,
    )

    return ViewConfig(
        filename=filename,
        view_type=view_type,
        display_type=display_type,
        case_name=_str("case_name"),
        mode_number=_int("mode_number", 1),
        window_number=_int("window_number", 0),
        render_delay=_float("render_delay", 0.0),
        azimuth=_float("azimuth", None) if data.get("azimuth") not in (None, "") else None,
        elevation=_float("elevation", None) if data.get("elevation") not in (None, "") else None,
        is_extruded=_bool("is_extruded", False),
        ui_automation_required=_bool("ui_automation_required", False),
        description=_str("description"),
    )


def _config_to_dict(cfg: ViewConfig) -> dict[str, object]:
    return {
        "filename": cfg.filename,
        "view_type": cfg.view_type.name,
        "display_type": cfg.display_type.name,
        "case_name": cfg.case_name,
        "mode_number": cfg.mode_number,
        "window_number": cfg.window_number,
        "render_delay": cfg.render_delay,
        "azimuth": cfg.azimuth,
        "elevation": cfg.elevation,
        "is_extruded": cfg.is_extruded,
        "ui_automation_required": cfg.ui_automation_required,
        "description": cfg.description,
    }


def serialize_plan(configs: list[ViewConfig]) -> list[dict[str, object]]:
    return [_config_to_dict(cfg) for cfg in configs]


def save_plan(configs: list[ViewConfig], output_path: str | Path = "capture_plan.json") -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as handle:
        json.dump(serialize_plan(configs), handle, ensure_ascii=False, indent=2)
    logger.info("Plan guardado en: %s", output_path)
    return output_path


def generate_sample_plan(output_path: str | Path = "capture_plan.json") -> Path:
    sample = [
        ViewConfig(
            filename="geometria_3d",
            view_type=ViewType.ISO_3D,
            display_type=DisplayType.GEOMETRY_ONLY,
            azimuth=225,
            elevation=30,
            description="Vista isométrica, solo geometría",
        ),
        ViewConfig(
            filename="geometria_planta",
            view_type=ViewType.PLAN_XY,
            display_type=DisplayType.GEOMETRY_ONLY,
            description="Vista en planta",
        ),
        ViewConfig(
            filename="geometria_elevacion_xz",
            view_type=ViewType.ELEV_XZ,
            display_type=DisplayType.GEOMETRY_ONLY,
            description="Elevación X-Z",
        ),
        ViewConfig(
            filename="carga_muerta_3d",
            view_type=ViewType.ISO_3D,
            display_type=DisplayType.LOAD_CASE,
            case_name="DEAD",
            azimuth=225,
            elevation=30,
            description="Carga muerta, vista 3D",
        ),
        ViewConfig(
            filename="carga_viva_3d",
            view_type=ViewType.ISO_3D,
            display_type=DisplayType.LOAD_CASE,
            case_name="LIVE",
            azimuth=225,
            elevation=30,
            description="Carga viva, vista 3D",
        ),
        ViewConfig(
            filename="sismo_x_planta",
            view_type=ViewType.PLAN_XY,
            display_type=DisplayType.LOAD_CASE,
            case_name="QUAKE-X",
            description="Sismo X, vista en planta",
        ),
        ViewConfig(
            filename="sismo_y_planta",
            view_type=ViewType.PLAN_XY,
            display_type=DisplayType.LOAD_CASE,
            case_name="QUAKE-Y",
            description="Sismo Y, vista en planta",
        ),
        ViewConfig(
            filename="modo_1_3d",
            view_type=ViewType.ISO_3D,
            display_type=DisplayType.MODE_SHAPE,
            case_name="MODAL",
            mode_number=1,
            render_delay=0.3,
            azimuth=225,
            elevation=30,
            description="Modo 1 de vibración",
        ),
        ViewConfig(
            filename="modo_2_3d",
            view_type=ViewType.ISO_3D,
            display_type=DisplayType.MODE_SHAPE,
            case_name="MODAL",
            mode_number=2,
            azimuth=225,
            elevation=30,
            description="Modo 2 de vibración",
        ),
        ViewConfig(
            filename="deformada_combo_env",
            view_type=ViewType.ISO_3D,
            display_type=DisplayType.DEFORMED,
            case_name="ENVELOPE",
            render_delay=0.5,
            azimuth=225,
            elevation=30,
            description="Forma deformada, envolvente",
        ),
    ]
    return save_plan(sample, output_path)


if __name__ == "__main__":
    generated = generate_sample_plan()
    print(f"Plan de ejemplo generado: {generated}")
