"""
=============================================================================
sap_imagenes.py
Captura automática de imágenes de SAP2000 v23 — integración con Excel
=============================================================================

DESCRIPCIÓN:
  Lee una tabla de configuración desde un archivo Excel por CLI
  o desde un workbook abierto vía xlwings, se conecta a SAP2000,
  aplica vistas, modos de display (modelo, cargas, resultados)
  y guarda las imágenes PNG.

MEJORAS DE SEGURIDAD:
  Usa exclusivamente el API de SAP2000 (OAPI). No mueve el mouse ni
  usa el teclado, garantizando que no se afecten otros programas.

REQUISITOS:
  pip install comtypes pywin32 Pillow openpyxl
=============================================================================
"""

import os
import sys
import time
import logging
import tempfile
import ctypes
import traceback
from pathlib import Path

# ---------------------------------------------------------------------------
# Importaciones
# ---------------------------------------------------------------------------
try:
    import openpyxl
except ImportError:
    raise ImportError("Instala openpyxl: pip install openpyxl")

comtypes = None
win32gui = None
win32con = None
Image = None
xw = None

def _ensure_comtypes():
    global comtypes
    if comtypes is not None: return comtypes
    try:
        import comtypes.client as comtypes_client
        import comtypes as _comtypes
        _comtypes.client = comtypes_client
        comtypes = _comtypes
    except ImportError:
        raise ImportError("Instala comtypes: pip install comtypes")
    return comtypes

def _ensure_pillow():
    global Image
    if Image is not None: return Image
    try:
        from PIL import Image as _Image
        Image = _Image
    except ImportError:
        raise ImportError("Instala Pillow: pip install Pillow")
    return Image

def _ensure_xlwings():
    global xw
    if xw is not None: return xw
    try:
        import xlwings as _xw
        xw = _xw
    except ImportError:
        raise ImportError("Instala xlwings: pip install xlwings")
    return xw

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------
SAP_DLL_PATH = r"C:\Program Files\Computers and Structures\SAP2000 23\SAP2000v1.dll"
PAUSA_TRAS_VISTA   = 0.5
PAUSA_TRAS_DISPLAY = 0.8

VISTAS_VALIDAS = {
    "PLANTA": (270, 90), "ELEV_X": (270, 0), "ELEV_Y": (0, 0),
    "ISO_NE": (225, 35), "ISO_NO": (315, 35), "ISO_SE": (135, 35), "ISO_SO": (45, 35),
    "CUSTOM": None
}

DISPLAY_MODELO, DISPLAY_CARGAS, DISPLAY_DEFORMADA, DISPLAY_FUERZAS, DISPLAY_DISENO = \
    "MODELO", "CARGAS", "DEFORMADA", "FUERZAS", "DISEÑO"

logging.basicConfig(level=logging.INFO, format="[%(asctime)s] %(levelname)s: %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("sap_imagenes")

SAFE_OUTPUT_ENVVAR = "SAP2000_ALLOW_UNSAFE_OUTPUT"

# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------
def valor_verdadero(valor) -> bool:
    v = str(valor or "").upper().strip()
    return v in ("SI", "SÍ", "YES", "1", "TRUE", "X")

def normalizar_crop(rect):
    if rect is None: return None
    try:
        iz, su, de, inf = [max(0.0, min(100.0, float(v))) for v in rect]
        if de <= iz or inf <= su: return None
        return (iz, su, de, inf)
    except: return None

def resolver_carpeta_salida(base_dir, conf, allow_unsafe=False):
    conf = str(conf or "Capturas_SAP").strip()
    if allow_unsafe:
        p = Path(conf)
        return str(p.resolve() if p.is_absolute() else (Path(base_dir) / p).resolve())
    if os.path.isabs(conf): raise ValueError("Ruta absoluta no permitida en modo seguro.")
    dest = (Path(base_dir).resolve() / conf).resolve()
    try:
        dest.relative_to(Path(base_dir).resolve())
        return str(dest)
    except: raise ValueError("La ruta de salida debe ser una subcarpeta del Excel.")

def permitir_salida_insegura(explicito=False):
    if explicito: return True
    return valor_verdadero(os.environ.get(SAFE_OUTPUT_ENVVAR, ""))

def mostrar_mensaje_portable(titulo, mensaje, icono="info"):
    if not getattr(sys, "frozen", False): return
    ic = {"info": 0x40, "warning": 0x30, "error": 0x10}.get(icono, 0x40)
    try: ctypes.windll.user32.MessageBoxW(0, str(mensaje), str(titulo), ic)
    except: pass

# ---------------------------------------------------------------------------
# Lógica de Captura
# ---------------------------------------------------------------------------
class SAP2000Conector:
    def __init__(self, dll_path=SAP_DLL_PATH):
        self.dll_path = dll_path
        self.sap_model = None

    def conectar(self):
        _ensure_comtypes()
        log.info("Iniciando conexión con SAP2000...")

        # Intento 1: Vía SAP2000v1.Helper (Recomendado)
        try:
            if os.path.exists(self.dll_path):
                try:
                    sap_gen = comtypes.client.GetModule(self.dll_path)
                    helper = comtypes.client.CreateObject("SAP2000v1.Helper").QueryInterface(sap_gen.cHelper)
                    sap_obj = helper.GetObject("CSI.SAP2000.API.SapObject")
                    self.sap_model = sap_obj.SapModel
                    log.info("Conexión vía Helper exitosa ✓")
                    return self
                except Exception as e:
                    log.warning(f"Fallo conexión vía Helper: {e}")
            else:
                log.warning(f"DLL no encontrado en {self.dll_path}. Intentando fallback...")
        except Exception as e:
            log.warning(f"Error durante inicialización de Helper: {e}")

        # Intento 2: Fallback directo a SapObject (si está registrado en el sistema)
        try:
            sap_obj = comtypes.client.GetActiveObject("CSI.SAP2000.API.SapObject")
            self.sap_model = sap_obj.SapModel
            log.info("Conexión vía GetActiveObject exitosa ✓")
            return self
        except Exception as e:
            log.warning(f"Fallo conexión directa: {e}")

        # Si llegamos aquí, no pudimos conectar
        msg = (
            "No se pudo conectar a SAP2000.\n\n"
            "Sugerencias:\n"
            "1. Asegúrate de que SAP2000 esté abierto con un modelo cargado.\n"
            "2. Si aparece 'Error al cargar la biblioteca de tipo/DLL', cierra SAP2000 y ejecuta "
            "como ADMINISTRADOR el archivo 'RegisterSAP2000.exe' que se encuentra en la carpeta de instalación de SAP2000.\n"
            f"3. Verifica que la ruta del DLL en la hoja CONFIG sea correcta: {self.dll_path}"
        )
        raise RuntimeError(msg)

class GestorVistas:
    def __init__(self, sap_model): self.sap_model = sap_model
    def set_vista(self, tipo, az=225, el=35):
        tipo = tipo.upper()
        angulos = VISTAS_VALIDAS.get(tipo)
        if angulos: az, el = angulos
        self.sap_model.View.SetView3D(az, el, 0)
        self.sap_model.View.RefreshView(0, True)
        time.sleep(PAUSA_TRAS_VISTA)

class GestorDisplay:
    def __init__(self, sap_model): self.sap_model = sap_model
    def set_modo(self, modo, item="", extrusion=False):
        try: self.sap_model.View.SetDisplayOptions(0, 11, extrusion) # Extrude
        except: pass

        modo = modo.upper()
        if modo == DISPLAY_CARGAS:
            # ShowLoadAssigns(Name, ItemType, ShowValues, ShowArrows, ...)
            # 2 = Frame, True = ShowValues
            self.sap_model.Display.ShowLoadAssigns(item, 2, True, False)
        elif modo == DISPLAY_DEFORMADA:
            # CaseComboName, ScaleFactor, ShowUnDeformed
            self.sap_model.Display.ShowDeformedShape(item, 0, True)
        elif modo == DISPLAY_FUERZAS:
            # Component 4 = Moment 3-3, 2 = Frame, 0 = Auto Scale, True = Show Values
            self.sap_model.Display.ShowForces(item, 4, 2, 0, True)
        elif modo == DISPLAY_DISENO:
            try:
                # Intentar Acero
                self.sap_model.DesignSteel.ShowResults(1, True)
            except:
                try:
                    # Intentar Concreto
                    self.sap_model.DesignConcrete.ShowResults(1, True)
                except:
                    log.warning("No se pudieron mostrar resultados de diseño (Acero/Concreto).")
        else:
            self.sap_model.View.RefreshWindow(0)
        time.sleep(PAUSA_TRAS_DISPLAY)

class GeneradorImagenes:
    def __init__(self, sap_model, salida, proy):
        self.sap_model, self.salida, self.proy = sap_model, salida, proy
        self.vistas = GestorVistas(sap_model)
        self.display = GestorDisplay(sap_model)
        os.makedirs(salida, exist_ok=True)

    def procesar(self, c):
        if not c.get("activo"):
            return {
                "ok": None, "estado": "inactiva", "archivo": "", "mensaje": "INACTIVO",
                "error_tipo": "", "contabiliza_error": False, "_fila": c.get("_fila"),
                "nombre_imagen": c.get("nombre_imagen", "")
            }
        try:
            self.vistas.set_vista(c["tipo_vista"], c["azimut"], c["elevacion"])
            self.display.set_modo(c["modo_display"], c["caso_combo"], c["extrusion"])
            nombre = f"{self.proy}_{c['nombre_imagen']}.png"
            ruta = os.path.join(self.salida, nombre)

            with tempfile.NamedTemporaryFile(suffix=".bmp", delete=False) as tmp: tmp_path = tmp.name
            try:
                # OAPI v23: SaveWindowToBMPFile(FileName)
                # NOTA: En algunas versiones es (WindowNumber, FileName, IncludeFrame)
                # Si falla con 1 argumento, probaremos con 3.
                try:
                    ret = self.sap_model.View.SaveWindowToBMPFile(tmp_path)
                except:
                    ret = self.sap_model.View.SaveWindowToBMPFile(0, tmp_path, False)

                if ret != 0: raise RuntimeError(f"Error OAPI SaveWindowToBMPFile: {ret}")

                img = _ensure_pillow().open(tmp_path)
                crop = normalizar_crop(c.get("crop"))
                if crop:
                    iw, ih = img.size
                    img = img.crop((int(iw*crop[0]/100), int(ih*crop[1]/100), int(iw*crop[2]/100), int(ih*crop[3]/100)))
                img.save(ruta, "PNG")
                log.info(f"  → Guardado: {nombre}")
                return {
                    "ok": True, "estado": "ok", "archivo": nombre, "mensaje": "OK",
                    "error_tipo": "", "contabiliza_error": False, "_fila": c["_fila"],
                    "nombre_imagen": c["nombre_imagen"]
                }
            finally:
                if os.path.exists(tmp_path):
                    try: os.remove(tmp_path)
                    except: pass
        except Exception as e:
            log.error(f"Error en fila {c.get('_fila')}: {e}")
            return {
                "ok": False, "estado": "error", "mensaje": str(e)[:120], "error_tipo": type(e).__name__,
                "contabiliza_error": True, "_fila": c["_fila"], "nombre_imagen": c["nombre_imagen"]
            }

# ---------------------------------------------------------------------------
# Funciones principales
# ---------------------------------------------------------------------------
def cargar_configuracion_desde_excel(ruta, allow_unsafe_output=False, proyecto_default="Modelo"):
    wb = openpyxl.load_workbook(ruta, data_only=True)
    try:
        sh_cfg, sh_cap = wb["CONFIG"], wb["CAPTURAS"]
        sap_dll = sh_cfg["B2"].value or SAP_DLL_PATH
        proy = sh_cfg["B3"].value or proyecto_default
        salida = resolver_carpeta_salida(os.path.dirname(ruta), sh_cfg["B4"].value, allow_unsafe_output)
        capturas = []
        for r in range(3, 1000):
            vals = [sh_cap.cell(row=r, column=c).value for c in range(1, 14)]
            if all(v in (None, "") for v in vals): break
            try:
                capturas.append({
                    "activo": valor_verdadero(vals[0]), "nombre_imagen": str(vals[1] or f"img_{r}"),
                    "tipo_vista": str(vals[2] or "ISO_NE").upper(), "azimut": float(vals[3] or 225),
                    "elevacion": float(vals[4] or 35), "modo_display": str(vals[5] or "MODELO").upper(),
                    "caso_combo": str(vals[6] or ""), "extrusion": valor_verdadero(vals[7]),
                    "ventana": str(vals[8] or "COMPLETA").upper(),
                    "crop": (vals[9], vals[10], vals[11], vals[12]) if str(vals[8]).upper() == "PARCIAL" else None,
                    "_fila": r
                })
            except Exception as e: log.warning(f"Fila {r} ignorada: {e}")
        return {"sap_dll_path": sap_dll, "nombre_proyecto": proy, "carpeta_salida": salida, "capturas": capturas}
    finally: wb.close()

def ejecutar_trabajo_capturas(config, sap_model=None, conectar_si_falta=True):
    conector = None
    if sap_model is None:
        if not conectar_si_falta: return {"ok": False, "stage": "conexion", "mensaje": "Sin sap_model", "resumen": {"ok": 0, "errores": 0, "inactivas": 0, "total": 0}, "resultados": []}
        try:
            conector = SAP2000Conector(config.get("sap_dll_path")).conectar()
            sap_model = conector.sap_model
        except Exception as e:
            return {"ok": False, "stage": "conexion", "mensaje": str(e), "resumen": {"ok": 0, "errores": 0, "inactivas": 0, "total": 0}, "resultados": []}

    gen = GeneradorImagenes(sap_model, config["carpeta_salida"], config["nombre_proyecto"])
    resultados = [gen.procesar(c) for c in config["capturas"]]
    ok = sum(1 for r in resultados if r.get("estado") == "ok")
    err = sum(1 for r in resultados if r.get("contabiliza_error"))
    inactivas = sum(1 for r in resultados if r.get("estado") == "inactiva")
    return {
        "ok": err == 0, "stage": "captura", "resultados": resultados,
        "resumen": {"ok": ok, "errores": err, "inactivas": inactivas, "total": len(resultados)},
        "conector": conector, "sap_model": sap_model, "carpeta_salida": config["carpeta_salida"],
        "mensaje": "OK" if err == 0 else "Errores parciales"
    }

def escribir_resultados_en_excel(ruta_excel, resultados):
    """Escribe resultados de vuelta al archivo Excel usando openpyxl."""
    try:
        wb = openpyxl.load_workbook(ruta_excel)
        sh_cap = wb["CAPTURAS"]
        for res in resultados:
            fila = res.get("_fila")
            if not fila: continue
            estado = res.get("estado", "error")
            txt = "✓ OK" if estado == "ok" else ("○ INACTIVA" if estado == "inactiva" else "✗ ERROR")
            sh_cap.cell(row=fila, column=14, value=txt) # Columna N
            sh_cap.cell(row=fila, column=15, value=res.get("archivo", "")) # Columna O
        wb.save(ruta_excel)
        log.info("Resultados escritos en el Excel ✓")
    except Exception as e: log.warning(f"No se pudo escribir resultados en Excel: {e}")

def escribir_resultados_xlwings(wb, resultados):
    """Escribe resultados usando xlwings (para integración con workbook abierto)."""
    try:
        sh_cap = wb.sheets["CAPTURAS"]
        for res in resultados:
            fila = res.get("_fila")
            if not fila: continue
            estado = res.get("estado", "error")
            txt = "✓ OK" if estado == "ok" else ("○ INACTIVA" if estado == "inactiva" else "✗ ERROR")
            color = (198, 239, 206) if estado == "ok" else ((217, 217, 217) if estado == "inactiva" else (255, 199, 206))
            sh_cap.range(f"N{fila}").value = txt
            sh_cap.range(f"N{fila}").color = color
            sh_cap.range(f"O{fila}").value = res.get("archivo", "")
    except Exception as e: log.warning(f"No se pudo escribir resultados vía xlwings: {e}")

def crear_excel_configuracion(ruta):
    wb = openpyxl.Workbook()
    sh_cfg = wb.active
    sh_cfg.title = "CONFIG"
    sh_cfg["A1"], sh_cfg["A2"], sh_cfg["B2"] = "CONFIGURACIÓN SAP2000 IMAGE CAPTURE", "Ruta DLL SAP2000", SAP_DLL_PATH
    sh_cfg["A3"], sh_cfg["B3"] = "Nombre del Proyecto", "MiProyecto"
    sh_cfg["A4"], sh_cfg["B4"] = "Subcarpeta de Salida", "Capturas_SAP"

    sh_cfg["A6"] = "SOLUCIÓN DE PROBLEMAS"
    sh_cfg["A7"] = "Si aparece 'Error al cargar la biblioteca de tipo/DLL':"
    sh_cfg["A8"] = "1. Cierra SAP2000."
    sh_cfg["A9"] = "2. Busca 'RegisterSAP2000.exe' en la carpeta de instalación de SAP2000."
    sh_cfg["A10"] = "3. Haz clic derecho y selecciona 'Ejecutar como administrador'."
    sh_cfg["A11"] = "4. Abre SAP2000 y vuelve a intentar."

    sh_cap = wb.create_sheet("CAPTURAS")
    headers = ["ACTIVO", "NOMBRE IMAGEN", "VISTA", "AZIMUT", "ELEVACIÓN", "DISPLAY", "CASO/COMBO", "EXTRUSIÓN", "VENTANA", "IZQ %", "SUP %", "DER %", "INF %", "ESTADO", "ARCHIVO"]
    for i, h in enumerate(headers, 1): sh_cap.cell(row=1, column=i, value=h).font = openpyxl.styles.Font(bold=True)
    ej = [
        ("SI", "Vista_General", "ISO_NE", 225, 35, "MODELO", "", "SI", "COMPLETA"),
        ("SI", "Cargas_Muertas", "ISO_NE", 225, 35, "CARGAS", "DEAD", "NO", "COMPLETA"),
        ("SI", "Deformada", "ISO_NE", 225, 35, "DEFORMADA", "DEAD", "NO", "COMPLETA"),
        ("SI", "Momentos", "ELEV_X", 270, 0, "FUERZAS", "COMB1", "NO", "COMPLETA"),
    ]
    for r, row in enumerate(ej, 2):
        for c, v in enumerate(row, 1): sh_cap.cell(row=r, column=c, value=v)
    wb.save(ruta)
    log.info(f"Excel creado: {ruta}")

def main_cli(argv=None):
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--crear-excel", metavar="R")
    parser.add_argument("--config", metavar="E")
    parser.add_argument("--allow-unsafe-output", action="store_true")
    args = parser.parse_args(argv)
    if args.crear_excel: crear_excel_configuracion(args.crear_excel)
    elif args.config:
        try:
            c = cargar_configuracion_desde_excel(args.config, args.allow_unsafe_output)
            r = ejecutar_trabajo_capturas(c)
            escribir_resultados_en_excel(args.config, r["resultados"])
            print(f"Resultado: {r['resumen']['ok']} OK, {r['resumen']['errores']} Error")
            return 0 if r['ok'] else 1
        except Exception as e:
            log.error(f"Error: {e}")
            mostrar_mensaje_portable("SAP2000 Capture", str(e), "error")
            return 1
    else:
        ruta = os.path.join(os.getcwd(), "SAP2000_Capturas.xlsx")
        if os.path.exists(ruta): return main_cli(["--config", ruta])
        crear_excel_configuracion(ruta)
        return 0

if __name__ == "__main__": sys.exit(main_cli())
