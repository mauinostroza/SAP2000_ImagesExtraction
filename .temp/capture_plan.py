"""
capture_plan.py
Lee el plan de capturas desde un archivo JSON o Excel (.xlsx).

Formato JSON:
[
  {
    "filename":     "vista_3d_muerta",
    "view_type":    "ISO_3D",
    "display_type": "LOAD_CASE",
    "case_name":    "DEAD",
    "description":  "Vista isométrica carga muerta"
  },
  {
    "filename":     "planta_sismo_x",
    "view_type":    "PLAN_XY",
    "display_type": "LOAD_CASE",
    "case_name":    "QUAKE-X"
  },
  {
    "filename":     "modo_1",
    "view_type":    "ISO_3D",
    "display_type": "MODE_SHAPE",
    "case_name":    "MODAL",
    "mode_number":  1
  }
]

Formato Excel (.xlsx) — columnas:
  filename | view_type | display_type | case_name | mode_number | render_delay | description

Dependencias para Excel: pip install openpyxl
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

from view_controller import ViewConfig, ViewType, DisplayType

logger = logging.getLogger(__name__)


def load_plan(path: str | Path) -> list[ViewConfig]:
    """Carga el plan de capturas desde JSON o Excel.

    Args:
        path: Ruta al archivo .json o .xlsx.

    Returns:
        Lista de ViewConfig listos para ejecutar en orden.

    Raises:
        ValueError: Si el formato no es reconocido o hay campos inválidos.
        FileNotFoundError: Si el archivo no existe.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Plan no encontrado: {path}")

    suffix = path.suffix.lower()
    if suffix == ".json":
        return _load_json(path)
    elif suffix in (".xlsx", ".xlsm"):
        return _load_excel(path)
    elif suffix == ".xls":
        raise ValueError(
            "Formato .xls no soportado por esta implementación; usa .xlsx o .json"
        )
    else:
        raise ValueError(f"Formato no soportado: '{suffix}'. Usa .json, .xlsx o .xlsm")


def _load_json(path: Path) -> list[ViewConfig]:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("El JSON debe ser una lista de objetos de captura")
    return [_dict_to_config(item, i) for i, item in enumerate(data)]


def _load_excel(path: Path) -> list[ViewConfig]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl no instalado. Ejecuta: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Detectar encabezados en fila 1
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    required = {"filename"}
    if not required.issubset(set(headers)):
        raise ValueError(f"El Excel debe tener al menos la columna 'filename'. Encontradas: {headers}")

    configs = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue  # fila vacía
        row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        try:
            configs.append(_dict_to_config(row_dict, row_num))
        except Exception as e:
            logger.warning(f"Fila {row_num} ignorada: {e}")

    return configs


def _dict_to_config(d: dict, index: int) -> ViewConfig:
    """Convierte un diccionario a ViewConfig, con valores por defecto."""

    def _str(key: str, default: str = "") -> str:
        v = d.get(key)
        return str(v).strip() if v is not None else default

    def _int(key: str, default: int = 0) -> int:
        v = d.get(key)
        try:
            return int(v) if v is not None else default
        except (TypeError, ValueError):
            return default

    def _float(key: str, default: float = 0.0) -> float:
        v = d.get(key)
        try:
            return float(v) if v is not None else default
        except (TypeError, ValueError):
            return default

    filename = _str("filename")
    if not filename:
        raise ValueError(f"Entrada {index}: 'filename' es obligatorio")

    # ViewType
    vt_str = _str("view_type", "ISO_3D").upper()
    try:
        view_type = ViewType[vt_str]
    except KeyError:
        valid = [e.name for e in ViewType]
        raise ValueError(f"view_type='{vt_str}' inválido. Opciones: {valid}")

    # DisplayType
    dt_str = _str("display_type", "GEOMETRY_ONLY").upper()
    try:
        display_type = DisplayType[dt_str]
    except KeyError:
        valid = [e.name for e in DisplayType]
        raise ValueError(f"display_type='{dt_str}' inválido. Opciones: {valid}")

    return ViewConfig(
        filename      = filename,
        view_type     = view_type,
        display_type  = display_type,
        case_name     = _str("case_name"),
        mode_number   = _int("mode_number", 1),
        window_number = _int("window_number", 0),
        render_delay  = _float("render_delay", 0.0),
        description   = _str("description"),
    )


def generate_sample_plan(output_path: str | Path = "capture_plan.json") -> Path:
    """Genera un archivo de ejemplo con capturas típicas de un proyecto."""
    sample = [
        {
            "filename":     "geometria_3d",
            "view_type":    "ISO_3D",
            "display_type": "GEOMETRY_ONLY",
            "description":  "Vista isométrica, solo geometría"
        },
        {
            "filename":     "geometria_planta",
            "view_type":    "PLAN_XY",
            "display_type": "GEOMETRY_ONLY",
            "description":  "Vista en planta"
        },
        {
            "filename":     "geometria_elevacion_xz",
            "view_type":    "ELEV_XZ",
            "display_type": "GEOMETRY_ONLY",
            "description":  "Elevación X-Z"
        },
        {
            "filename":     "carga_muerta_3d",
            "view_type":    "ISO_3D",
            "display_type": "LOAD_CASE",
            "case_name":    "DEAD",
            "description":  "Carga muerta, vista 3D"
        },
        {
            "filename":     "carga_viva_3d",
            "view_type":    "ISO_3D",
            "display_type": "LOAD_CASE",
            "case_name":    "LIVE",
            "description":  "Carga viva, vista 3D"
        },
        {
            "filename":     "sismo_x_planta",
            "view_type":    "PLAN_XY",
            "display_type": "LOAD_CASE",
            "case_name":    "QUAKE-X",
            "description":  "Sismo X, vista en planta"
        },
        {
            "filename":     "sismo_y_planta",
            "view_type":    "PLAN_XY",
            "display_type": "LOAD_CASE",
            "case_name":    "QUAKE-Y",
            "description":  "Sismo Y, vista en planta"
        },
        {
            "filename":     "modo_1_3d",
            "view_type":    "ISO_3D",
            "display_type": "MODE_SHAPE",
            "case_name":    "MODAL",
            "mode_number":  1,
            "render_delay": 0.3,
            "description":  "Modo 1 de vibración"
        },
        {
            "filename":     "modo_2_3d",
            "view_type":    "ISO_3D",
            "display_type": "MODE_SHAPE",
            "case_name":    "MODAL",
            "mode_number":  2,
            "description":  "Modo 2 de vibración"
        },
        {
            "filename":     "deformada_combo_env",
            "view_type":    "ISO_3D",
            "display_type": "DEFORMED",
            "case_name":    "ENVELOPE",
            "render_delay": 0.5,
            "description":  "Forma deformada, envolvente"
        },
    ]
    output_path = Path(output_path)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(sample, f, ensure_ascii=False, indent=2)
    logger.info(f"Plan de ejemplo generado en: {output_path}")
    return output_path


if __name__ == "__main__":
    p = generate_sample_plan()
    print(f"Plan de ejemplo generado: {p}")
    configs = load_plan(p)
    print(f"{len(configs)} capturas cargadas:")
    for c in configs:
        print(f"  {c.filename:30s} {c.view_type.name:10s} {c.display_type.name:16s} {c.case_name}")
